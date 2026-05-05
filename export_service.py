from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook
import pandas as pd

from data_fetcher import StockDataFetcher


def _format_daily_sheet(daily_df: pd.DataFrame) -> pd.DataFrame:
    if daily_df.empty:
        return daily_df

    formatted = daily_df.copy()
    formatted["Date"] = pd.to_datetime(formatted["Date"])
    return formatted


def _format_revenue_sheet(revenue_df: pd.DataFrame) -> pd.DataFrame:
    if revenue_df.empty:
        return revenue_df

    formatted = revenue_df.copy()
    formatted["Month"] = pd.to_datetime(formatted["Month"])
    return formatted


def build_stock_export_workbook(symbol_or_code: str) -> tuple[BytesIO, str, str]:
    with StockDataFetcher() as fetcher:
        meta = fetcher.resolve_stock(symbol_or_code)

        price_df = fetcher.fetch_price_history(meta)
        trading_dates = price_df["Date"].tolist()

        institutional_df = fetcher.fetch_institutional_daily(meta, trading_dates)
        margin_df = fetcher.fetch_margin_daily(meta, trading_dates)
        revenue_df = fetcher.fetch_monthly_revenue(meta, start_year=2023)
        financial_df = fetcher.fetch_quarterly_financials(meta)

        daily_df = fetcher.merge_daily_frames(price_df, institutional_df, margin_df)
        summary_df = fetcher.build_strategy_summary(meta, daily_df, revenue_df, financial_df)

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        _format_daily_sheet(daily_df).to_excel(writer, sheet_name="Price_History", index=False)
        _format_revenue_sheet(revenue_df).to_excel(writer, sheet_name="Monthly_Revenue", index=False)
        financial_df.to_excel(writer, sheet_name="Quarterly_Financials", index=False)
        summary_df.to_excel(writer, sheet_name="Strategy_Summary", index=False)

    buffer.seek(0)
    file_name = f"{meta.code}_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return buffer, file_name, meta.display_name


def inspect_stock_export(symbol_or_code: str, sample_rows: int = 3) -> dict[str, object]:
    buffer, file_name, display_name = build_stock_export_workbook(symbol_or_code)
    workbook = load_workbook(buffer, data_only=True)

    sheets: dict[str, dict[str, object]] = {}
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        header = [str(value) if value is not None else "" for value in (rows[0] if rows else [])]
        preview = [list(row) for row in rows[1 : 1 + sample_rows]]
        sheets[sheet_name] = {
            "rows": max(worksheet.max_row - 1, 0),
            "columns": header,
            "preview": preview,
        }

    return {
        "file_name": file_name,
        "display_name": display_name,
        "sheets": sheets,
    }