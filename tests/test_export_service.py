from __future__ import annotations

import io
from io import BytesIO
import importlib
from pathlib import Path
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import load_workbook
import pandas as pd

import export_service


EXPECTED_SHEETS = {
    "Price_History",
    "Monthly_Revenue",
    "Quarterly_Financials",
    "Strategy_Summary",
}


def _build_sample_export_buffer() -> BytesIO:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        pd.DataFrame(
            [
                {"Date": "2026-06-18", "Close": 100.0, "Volume": 1234},
                {"Date": "2026-06-19", "Close": 101.5, "Volume": 2345},
            ]
        ).to_excel(writer, sheet_name="Price_History", index=False)
        pd.DataFrame(
            [
                {"Month": "2026-05-01", "Revenue": 1000000},
                {"Month": "2026-06-01", "Revenue": 1200000},
            ]
        ).to_excel(writer, sheet_name="Monthly_Revenue", index=False)
        pd.DataFrame(
            [
                {"Quarter": "2026Q1", "EPS": 1.2},
            ]
        ).to_excel(writer, sheet_name="Quarterly_Financials", index=False)
        pd.DataFrame(
            [
                {"Metric": "Close", "Value": 101.5},
            ]
        ).to_excel(writer, sheet_name="Strategy_Summary", index=False)
    buffer.seek(0)
    return buffer


class ExportServiceInspectionTests(unittest.TestCase):
    def test_inspect_stock_export_workbook_reads_existing_buffer(self) -> None:
        buffer = _build_sample_export_buffer()

        summary = export_service.inspect_stock_export_workbook(
            buffer,
            "2330_export.xlsx",
            "2330 台積電",
            sample_rows=1,
        )

        self.assertEqual(summary["file_name"], "2330_export.xlsx")
        self.assertEqual(summary["display_name"], "2330 台積電")
        self.assertEqual(set(summary["sheets"].keys()), EXPECTED_SHEETS)
        self.assertEqual(summary["sheets"]["Price_History"]["rows"], 2)
        self.assertEqual(summary["sheets"]["Price_History"]["columns"], ["Date", "Close", "Volume"])
        self.assertEqual(len(summary["sheets"]["Price_History"]["preview"]), 1)

    def test_inspect_stock_export_workbook_leaves_buffer_reusable(self) -> None:
        buffer = _build_sample_export_buffer()

        export_service.inspect_stock_export_workbook(
            buffer,
            "2330_export.xlsx",
            "2330 台積電",
        )

        workbook = load_workbook(buffer, data_only=True)
        try:
            self.assertEqual(set(workbook.sheetnames), EXPECTED_SHEETS)
        finally:
            workbook.close()

    def test_inspect_stock_export_wrapper_builds_workbook_once(self) -> None:
        buffer = _build_sample_export_buffer()

        with patch.object(
            export_service,
            "build_stock_export_workbook",
            return_value=(buffer, "2330_export.xlsx", "2330 台積電"),
        ) as build_workbook:
            summary = export_service.inspect_stock_export("2330", sample_rows=2)

        build_workbook.assert_called_once_with("2330")
        self.assertEqual(summary["sheets"]["Monthly_Revenue"]["rows"], 2)

    def test_local_export_verifier_reuses_one_workbook_for_preview_and_save(self) -> None:
        export_cli = importlib.import_module("test")
        buffer = _build_sample_export_buffer()

        with tempfile.TemporaryDirectory() as temp_dir:
            save_path = Path(temp_dir) / "2330_export.xlsx"
            with (
                patch.object(
                    export_cli,
                    "build_stock_export_workbook",
                    return_value=(buffer, "2330_export.xlsx", "2330 台積電"),
                ) as build_workbook,
                patch.object(
                    export_cli,
                    "inspect_stock_export_workbook",
                    wraps=export_cli.inspect_stock_export_workbook,
                ) as inspect_workbook,
                patch(
                    "sys.argv",
                    ["test.py", "2330", "--save", str(save_path), "--preview-rows", "1"],
                ),
                patch("sys.stdout", new_callable=io.StringIO),
            ):
                export_cli.main()

            build_workbook.assert_called_once_with("2330")
            inspect_workbook.assert_called_once()
            self.assertGreater(save_path.stat().st_size, 1000)


if __name__ == "__main__":
    unittest.main()
